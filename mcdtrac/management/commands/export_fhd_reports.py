from django.core.management.base import BaseCommand, CommandError
from django.db import connection, transaction
from django.conf import settings
import pprint
import datetime
import dateutil
import openpyxl
import os
from mcdtrac.utils import dictfetchall, XLS_DIR
from optparse import make_option
from rapidsms.contrib.locations.models import Location

class Command(BaseCommand):
    """export excels of the latest week/quarterly report"""
    help = 'Creates quarterly/weekly spreadsheets for download.'
    pp = pprint.PrettyPrinter(indent=4)  # debug
    base_width = 10.47
    individual_col_widths = {
                'A': base_width * 1.5,
                'D': base_width * 1.5,
                'E': base_width * 2,
                'F': base_width * 3
    }
    # optparse class variables
    debug_fhd = False
    fhd_districts = []
    start_date = None
    end_date = None
    custom_range = False

    option_list = BaseCommand.option_list + (
            make_option(
                '--debug',
                action='store_true',
                dest='debug_fhd',
                default=False,
                help='Debug FHD CMD'
            ),
            make_option(
                '-d',
                '--district',
                action='append',
                help='Add this district to the list of districts to export.'
            ),
            make_option(
                '-s',
                '--start-date',
                help='Collect data starting on this date. Default: start of this quarter.'
            ),
            make_option(
                '-e',
                '--end-date',
                help='Collect data ending on this date. Default: today.'
            )
        )

    def get_districts(self):
        """rows of district objects with FHD data."""
        cursor = connection.cursor()
        sql = """SELECT l.lft,
                   l.id,
                   l.name AS District ,
                   l.rght
            FROM fhd_stats_mview f ,
                 locations_location l
            WHERE f.lft <= l.rght
                AND f.lft >= l.lft
                AND l.id IN
                    (SELECT "locations_location"."id"
                     FROM "locations_location"
                     WHERE ("locations_location"."lft" <= 15257
                            AND "locations_location"."lft" >= 2
                            AND "locations_location"."tree_id" = 1
                            AND "locations_location"."type_id" = 'district'))
            GROUP BY l.lft,
                     l.id,
                     l.name,
                     l.rght"""
        cursor.execute(sql)
        rows = dictfetchall(cursor)
        transaction.commit_unless_managed()
        return rows

    def get_district_by_name(self, district):
        """parse command line district name into a dict"""
        try:
            d = Location.tree.root_nodes(
                )[0].get_descendants(
                ).get(
                    type='district',
                    name__iexact=district
                )
        except Exception, e:
            raise CommandError('District "{0}" returned error: {1}'.format(district, e))
        else:
            return {
                'lft': d.lft,
                'id': d.id,
                'district': d.name,
                'rght': d.rght
            }

    def generate_sql(self, sub_level=False, location_id=None):
        """
        generate SQL used by the handle() below as follows:

        if sub_level = False:
            generate the strings suitable for a global level worksheets
        else:
            generate the sql suitable for a particular location specified
            as the <location_id> option. Raise CommandError if no <location_id>
        """
        if sub_level:
            if not (location_id and isinstance(location_id, (int, long))):
                raise CommandError("If you want a district give me a location_id")
            else:
                grp_sql_title = 'SELECT f.facility AS "Facility"'
                sql_id = 'l.id = {0}'.format(int(location_id))
                grp_sql_name = 'f.facility'
                grp_extra_cols = ''
        else:
            grp_sql_title = 'SELECT l.name AS "District"'
            sql_id = """l.id in (SELECT "locations_location"."id"
                FROM "locations_location"
                WHERE ("locations_location"."lft" <= 15257
                       AND "locations_location"."lft" >= 2
                       AND "locations_location"."tree_id" = 1
                       AND "locations_location"."type_id" = 'district'))"""
            grp_sql_name = 'l.name'
            grp_extra_cols = """'' AS "Zonal Office",
            '' AS "Expected POW Outreaches",
            COUNT(pow_name) AS "POW Outreach Data",
            '' AS "POW reporting rates",
            '' AS "Responsible DPO","""

        grouped_sql = """{0},
                COUNT(DISTINCT xformreportsubmission_id) AS "Entries",
                {3}
                SUM(dpt_male) AS "DPT (M)",
                SUM(dpt_female) AS "DPT (F)",
                SUM(vacm_male) AS "Measles (M)",
                SUM(vacm_female) AS "Measles (F)",
                SUM(vita_male1) AS "Vitamin A Males (6-11m)",
                SUM(vita_female1) AS "Vitamin A Females (6-11m)",
                SUM(vita_male2) AS "Vitamin A Males (12-59m)",
                SUM(vita_female2) AS "Vitamin A Females (12-59m)",
                SUM(pcv_pcv1_m) AS "PCV1 (M)",
                SUM(pcv_pcv1_f) AS "PCV1 (F)",
                SUM(pcv_pcv2_m) AS "PCV2 (M)",
                SUM(pcv_pcv2_f) AS "PCV2 (F)",
                SUM(pcv_pcv3_m) AS "PCV3 (M)",
                SUM(pcv_pcv3_f) AS "PCV3 (F)",
                SUM(dorm_male_1to4) AS "Dorm (M) (1-4y)",
                SUM(dorm_female_1to4) AS "Dorm (F) (1-4y)",
                SUM(dorm_male_5to14) AS "Dorm (M) (5-14y)",
                SUM(dorm_female_5to14) AS "Dorm (F) (5-14y)",
                SUM(ryg_red_male) AS "Red MUAC (M)",
                SUM(ryg_red_female) AS "Red MUAC (F)",
                SUM(ryg_yellow_male) AS "Yellow MUAC (M)",
                SUM(ryg_yellow_female) AS "Yellow MUAC (F)",
                SUM(ryg_green_male) AS "Green MUAC (M)",
                SUM(ryg_green_female) AS "Green MUAC (F)",
                SUM(tetn_dose1) AS "TETN Dose 1",
                SUM(tetn_dose2) AS "TETN Dose 2",
                SUM(tetn_dose3) AS "TETN Dose 3",
                SUM(tetn_dose4) AS "TETN Dose 4",
                SUM(tetn_dose5) AS "TETN Dose 5",
                SUM(npet_dose1) AS "NPET Dose 1",
                SUM(npet_dose2) AS "NPET Dose 2",
                SUM(npet_dose3) AS "NPET Dose 3",
                SUM(npet_dose4) AS "NPET Dose 4",
                SUM(npet_dose5) AS "NPET Dose 5",
                SUM(ancp_anc1) AS "# ANC1",
                SUM(ancp_anc4) AS "# ANC4",
                SUM(ancp_ipt1) AS "# IPT1",
                SUM(ancp_ipt2) AS "# IPT2",
                SUM(heid_male) AS "HEID Children (M)",
                SUM(heid_female) AS "HEID Children (F)",
                SUM(heid_tot) AS "HEID Children (Total)",
                SUM(heid_women) AS "HEID Women Tested",
                SUM(breg_male) AS "Birth Registration (M)",
                SUM(breg_female) AS "Birth Registration (F)",
                SUM(bpbs_bpm) AS "Blood pressure (M)",
                SUM(bpbs_bpf) AS "Blood pressure (F)",
                SUM(bpbs_bsm) AS "Blood sugar (M)",
                SUM(bpbs_bsf) AS "Blood sugar (F)"
            FROM fhd_stats_mview f,
                 locations_location l
            WHERE f.has_errors = FALSE
                AND f.created >= %s
                AND f.created <= %s
                AND f.lft <= l.rght
                AND f.lft >= l.lft
                AND {1}
            GROUP BY l.lft,
                     l.id,
                     {2},
                     l.rght""".format(grp_sql_title, sql_id, grp_sql_name, grp_extra_cols)

        individual_sql = """SELECT f.submission_id,
                    f.created::date AS "Date",
                    l.name AS district,
                    f.facility AS "Name of Facility",
                    f.reporting_name AS "Name of Reporter",
                    f.phone AS "Phone Number of Reporter",
                    f.has_errors AS "Invalid",
                    f.pow_district_name AS "POW District",
                    f.pow_code AS "POW Code",
                    f.pow_name AS "POW Name",
                    f.dpt_male AS "DPT (M)",
                    f.dpt_female AS "DPT (F)",
                    f.vacm_male AS "Measles (M)",
                    f.vacm_female AS "Measles (F)",
                    f.vita_male1 AS "Vitamin A Males (6-11m)",
                    f.vita_female1 AS "Vitamin A Females (6-11m)",
                    f.vita_male2 AS "Vitamin A Males (12-59m)",
                    f.vita_female2 AS "Vitamin A Females (12-59m)",
                    f.pcv_pcv1_m AS "PCV1 (M)",
                    f.pcv_pcv1_f AS "PCV1 (F)",
                    f.pcv_pcv2_m AS "PCV2 (M)",
                    f.pcv_pcv2_f AS "PCV2 (F)",
                    f.pcv_pcv3_m AS "PCV3 (M)",
                    f.pcv_pcv3_f AS "PCV3 (F)",
                    f.dorm_male_1to4 AS "Dorm (M) (1-4y)",
                    f.dorm_female_1to4 AS "Dorm (F) (1-4y)",
                    f.dorm_male_5to14 AS "Dorm (M) (5-14y)",
                    f.dorm_female_5to14 AS "Dorm (F) (5-14y)",
                    f.ryg_red_male AS "Red MUAC (M)",
                    f.ryg_red_female AS "Red MUAC (F)",
                    f.ryg_yellow_male AS "Yellow MUAC (M)",
                    f.ryg_yellow_female AS "Yellow MUAC (F)",
                    f.ryg_green_male AS "Green MUAC (M)",
                    f.ryg_green_female AS "Green MUAC (F)",
                    f.tetn_dose1 AS "TETN of Dose 1",
                    f.tetn_dose2 AS "TETN of Dose 2",
                    f.tetn_dose3 AS "TETN of Dose 3",
                    f.tetn_dose4 AS "TETN of Dose 4",
                    f.tetn_dose5 AS "TETN of Dose 5",
                    f.npet_dose1 AS "NPET Dose 1",
                    f.npet_dose2 AS "NPET Dose 2",
                    f.npet_dose3 AS "NPET Dose 3",
                    f.npet_dose4 AS "NPET Dose 4",
                    f.npet_dose5 AS "NPET Dose 5",
                    f.ancp_anc1 AS "# ANC1",
                    f.ancp_anc4 AS "# ANC4",
                    f.ancp_ipt1 AS "# IPT1",
                    f.ancp_ipt2 AS "# IPT2",
                    f.heid_male AS "HEID Children (M)",
                    f.heid_female AS "HEID Children (F)",
                    f.heid_tot AS "HEID Children (Total)",
                    f.heid_women AS "HEID Women Tested",
                    f.breg_male AS "Birth Registration (M)",
                    f.breg_female AS "Birth Registration (F)",
                    f.bpbs_bpm AS "Blood pressure (M)",
                    f.bpbs_bpf AS "Blood pressure (F)",
                    f.bpbs_bsm AS "Blood sugar (M)",
                    f.bpbs_bsf AS "Blood sugar (F)",
                    f.worm_male AS "** WORM (M)",
                    f.worm_female AS "** WORM (F)",
                    f.redm_number AS "** REDM",
                    f.tet_dose2 AS "** TET dose2",
                    f.tet_dose3 AS "** TET dose3",
                    f.tet_dose4 AS "** TET dose4",
                    f.tet_dose5 AS "** TET dose5",
                    f.anc_number AS "** ANC >4 ",
                    f.eid_male AS "** EID (M)",
                    f.eid_female AS "** EID (F)"
            FROM fhd_stats_mview f ,
                 locations_location l
            WHERE f.created >= %s
                AND f.created <= %s
                AND f.lft <= l.rght
                AND f.lft >= l.lft
                AND {0}""".format(sql_id)

        return (grouped_sql, individual_sql)

    def populate_worksheet(self, ws=None, title='', sql_list=[], col_widths={}):

        if not (ws and title and sql_list):
            raise CommandError('Got incomplete arguments')

        cursor = connection.cursor()
        ws.title = title
        ws.show_gridlines = False  # doesn't seem to work
        sql_str = sql_list.pop(0)

        if self.debug_fhd:
            self.stdout.write('DEBUG::SQL:\n{0}\n\n\tstart:{1}\tend:{2}\n\n'.format(sql_str ,sql_list[0], sql_list[1]))


        cursor.execute(sql_str, sql_list)
        headers = [col[0] for col in cursor.description]
        rows = dictfetchall(cursor)
        transaction.commit_unless_managed()

        col = 0
        row = 0
        for h in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = h
            col += 1

        col = 0
        row = 1
        for r in rows:
            for key in r:
                ws.cell(row=row, column=col).value = r[key]
                col += 1
            row += 1
            col = 0

        ws.row_dimensions[1].height = 42.0
        for c in ws.column_dimensions:
            ws.column_dimensions[c].width = self.base_width

        if col_widths:
            for c in col_widths:
                ws.column_dimensions[c].width = col_widths[c]

        for cell in ws.rows[0]:
            cell.style.alignment.vertical = openpyxl.style.Alignment.VERTICAL_TOP
            cell.style.alignment.wrap_text = True
            cell.style.font.bold = True
            cell.style.fill.fill_type = openpyxl.style.Fill.FILL_SOLID
            cell.style.fill.start_color.index = 'FFE6E6E6'
            #todo: add borders
        return ws

    def generate_fpath(self, subdir='.', prefix='', suffix=''):
        q_str = 'Q' + str((self.start_date.month - 1) // 3 + 1)
        y_str = str(self.start_date.year)

        if self.custom_range:
            xls_fname = '{2}fhd_stats_{0}_{1}{3}.xlsx'.format(
                self.start_date.strftime('%F'),
                self.end_date.strftime('%F'),
                prefix.lower(),
                suffix.lower())
        else:
            xls_fname = '{2}fhd_stats-{0}_{1}{3}.xlsx'.format(
                y_str,
                q_str,
                prefix.lower(),
                suffix.lower())

        xls_fpath = os.path.join(
            os.path.split(settings.MTRACK_ROOT)[0],
            XLS_DIR,
            '{0}/{1}'.format(subdir.lower(), y_str),
            xls_fname
        )
        return os.path.abspath(xls_fpath)

    def handle(self, *args, **options):
        # TODO: add function to create xls fpath based on district name
        # or 'uganda' for country level xls.
        self.debug_fhd = options['debug_fhd']
        quarter_months = ['01', '04', '07', '10']

        if self.debug_fhd:
            self.stdout.write('got options: ' + self.pp.pformat(options) + '\n')

        if options['start_date']:
            self.custom_range = True
            self.start_date = dateutil.parser.parse(options['start_date']).date()
        else:
            self.start_date = dateutil.parser.parse(
                str(datetime.date.today().year) + '-' +
                str(quarter_months[(datetime.date.today().month - 1) // 3]) + '-' +
                '01'
            ).date()

        if options['end_date']:
            self.custom_range = True
            self.end_date = dateutil.parser.parse(options['end_date']).date()
        else:
            self.end_date = datetime.date.today()

        if options['district']:
            for d in options['district']:
                self.fhd_districts.append(self.get_district_by_name(d))
        else:
            self.fhd_districts = self.get_districts()

        if self.debug_fhd:
            self.stdout.write(
                'DEBUG: doing districts: ' +
                self.pp.pformat(self.fhd_districts) +
                '\n')

        for fhd_dist in self.fhd_districts:
            wb = openpyxl.Workbook()
            xls_fpath = self.generate_fpath(
                            fhd_dist['district'],
                            fhd_dist['district'] + '-')
            try:
                os.makedirs(os.path.dirname(xls_fpath))
            except OSError:
                pass
            if self.debug_fhd:
                self.stdout.write(
                    'DEBUG: Writing spreadsheet to: "{0}"\n'.format(xls_fpath)
                )
            grouped_sql, individual_sql = self.generate_sql(
                                            sub_level=True,
                                            location_id = fhd_dist['id'])
            self.stdout.write(':: facility entries for {0}\n'.format(
                                fhd_dist['district']))
            self.populate_worksheet(
                ws=wb.get_active_sheet(),
                title="Facility Summaries",
                sql_list=[grouped_sql, self.start_date, self.end_date]
            )
            self.stdout.write(':: individual entries for {0}\n'.format(
                                fhd_dist['district']))
            self.populate_worksheet(
                ws=wb.create_sheet(),
                title="Individual Enetries",
                sql_list=[individual_sql, self.start_date, self.end_date],
                col_widths=self.individual_col_widths
            )
            if self.debug_fhd:
                self.stdout.write(
                    'DEBUG: Workbook is: ' +
                    self.pp.pformat(wb.worksheets) + '\n')

            wb.save(xls_fpath)

        wb = openpyxl.Workbook()
        xls_fpath = self.generate_fpath('uganda')

        try:
            os.makedirs(os.path.dirname(xls_fpath))
        except OSError:
            pass
        if self.debug_fhd:
            self.stdout.write(
                'DEBUG: Writing spreadsheet to: "{0}"\n'.format(xls_fpath)
            )
        grouped_sql, individual_sql = self.generate_sql()
        self.stdout.write(':: Generating country level district summaries.\n')
        self.populate_worksheet(
            ws=wb.get_active_sheet(),
            title="District Summaries",
            sql_list=[grouped_sql, self.start_date, self.end_date])
        self.stdout.write(':: Generating country level individual entries.\n')
        self.populate_worksheet(
            ws=wb.create_sheet(),
            title="Individual Entries",
            sql_list=[individual_sql, self.start_date, self.end_date],
            col_widths=self.individual_col_widths)
        if self.debug_fhd:
            self.stdout.write(
                    'DEBUG: Workbook is: ' +
                    self.pp.pformat(wb.worksheets) + '\n'
                )
        wb.save(xls_fpath)
